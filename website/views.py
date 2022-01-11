from flask import Blueprint, render_template, request, flash, jsonify
from flask_login import login_required, current_user
from .models import Note
from .models import User
from .models import SemesterGrade
from . import db
import json
import openpyxl

views = Blueprint('views', __name__)

def courseGrades(startRow, startCol, sheet_obj):
    cell_obj = sheet_obj.cell(row = startRow, column = startCol)
    courseName = sheet_obj.cell(row = 3, column = 3).value
    while(cell_obj.value != None):
        # get user and add info
        curUser = User.query.filter_by(matriculationNumber=cell_obj.value).first()
        if curUser != None:
            allNotes = curUser.notes
            check = False
            for n in allNotes:
                if n.courseName == courseName:
                    n.totalScore = sheet_obj.cell(row = startRow, column = startCol + 1).value
                    n.letterGrade = sheet_obj.cell(row = startRow, column = startCol + 2).value
                    n.exam = sheet_obj.cell(row = startRow, column = startCol + 12).value
                    n.ca = sheet_obj.cell(row = startRow, column = startCol + 13).value
                    db.session.commit() 
                    check = True
                    break
            if check == False:
                ts = sheet_obj.cell(row = startRow, column = startCol + 1).value
                lg = sheet_obj.cell(row = startRow, column = startCol + 2).value
                e = sheet_obj.cell(row = startRow, column = startCol + 12).value
                c = sheet_obj.cell(row = startRow, column = startCol + 13).value
                new_note = Note(courseName=courseName,data=cell_obj.value,totalScore=ts,letterGrade=lg,exam=e,ca=c,user_id=curUser.id)
                db.session.add(new_note)
                db.session.commit()        
        startRow += 1
        cell_obj = sheet_obj.cell(row = startRow, column = startCol)

@views.route('/', methods=['GET', 'POST'])
@login_required
def home1():
    check = False
    if current_user.matriculationNumber == "admin":
        check = True
    return render_template("see.html", user=current_user, isAdmin=check, users=User.query.all())
#    return render_template("base.html", user=current_user)
    

@views.route('/<path:pars>', methods=['GET', 'POST'])
@login_required
def home(pars):
    print(current_user.matriculationNumber)
    print(pars)
    print(current_user.matriculationNumber != pars  and current_user.matriculationNumber != "admin")
    if current_user.matriculationNumber != pars and current_user.matriculationNumber != "admin":
        #return render_template("base.html", user=current_user)
        flash('You do not have access to that page.', category='error')
        return render_template("home.html", user=current_user, isAdmin=False)


    if request.method == 'POST':
        ####
        file = request.files.get('file')
        wb_obj = openpyxl.load_workbook(file, data_only=True)
        sheet_obj = wb_obj.active
        test = sheet_obj.cell(row = 1, column = 1).value
        if test == "UNIVERSITY OF PORT HARCOURT COURSE EXAMINATION MARK SHEET":
            startRow = 7
            startCol = 2
            courseGrades(startRow, startCol, sheet_obj)
            flash('Course Grade added!', category='success')
        test = sheet_obj.cell(row = 3, column = 1).value
        if test == "COURSE RECORD SHEET FOR B.ENG. (GAS ENGINEERING) DEGREE PROGRAMME":
            startCol = 1
            checkRows = [10, 26, 62, 78, 113, 129, 164, 180, 215, 231, 266, 282, 317, 330]

            cell_obj = sheet_obj.cell(row = 7, column = 3)
            curUser = User.query.filter_by(matriculationNumber=cell_obj.value).first()
            for i in range(0, len(checkRows), 1):
                startRow = checkRows[i]
                cell_obj = sheet_obj.cell(row = startRow, column = startCol)
                semesterTitle = cell_obj.value
                print(333)
                gpa = sheet_obj.cell(row = startRow + 14, column = startCol + 6).value
                tcu = sheet_obj.cell(row = startRow + 12, column = startCol + 3).value
                
                if startRow in [10, 26]:
                    year = 1
                elif startRow in [62, 78]:
                    year = 2
                elif startRow in [113, 129]:
                    year = 3
                elif startRow in [164, 180]:
                    year = 4
                elif startRow in [215, 231]:
                    year = 5
                elif startRow in [266, 282]:
                    year = 6
                else:
                    year = 7

                if "FIRST" in semesterTitle:
                    semester = 1
                else:
                    semester = 2
                curCol = startCol + 1
                curRow = startRow + 2
                curCell = sheet_obj.cell(row = curRow, column = curCol)
                while(curCell.value != None):
                    print(4444)
                    """
                        courseName = db.Column(db.String(100))
                            courseName = db.Column(db.String(100))
                            courseTitle = db.Column(db.String(1000))
                            year = db.Column(db.Integer)
                            semester = db.Column(db.String(10))
                            score = db.Column(db.Integer)
                            grade = db.Column(db.String(10))
                            gpa = db.Column(db.String(10))
                            tcu = db.Column(db.String(10))
                            ####
                            date = db.Column(db.DateTime(timezone=True), default=func.now())
                            user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
                        """
                        
                    cn = sheet_obj.cell(row = curRow, column = curCol).value
                    ct = sheet_obj.cell(row = curRow, column = curCol + 1).value
                    s = sheet_obj.cell(row = curRow, column = curCol + 3).value
                    g = sheet_obj.cell(row = curRow, column = curCol + 4).value

                    check = False
                    for sG in curUser.semesterGrades:
                        if sG.courseName == cn:
                            sG.courseName = cn
                            sG.year = year
                            sG.courseTitle=ct
                            sG.semester=semester
                            sG.score=s
                            sG.grade=g
                            sG.gpa=gpa
                            sG.tcu=tcu
                            db.session.commit() 
                            check = True
                            break
                    if check == False:
                        new_sg = SemesterGrade(courseName=cn,year=year,courseTitle=ct,semester=semester,score=s,grade=g,gpa=gpa,tcu=tcu,user_id=curUser.id)
                        db.session.add(new_sg)
                        db.session.commit() 
                    curRow += 1
                    curCell = sheet_obj.cell(row = curRow, column = curCol)
                    # update the start roe and startcol and semestertitle
                    """
                    startRow += 16
                    cell_obj = sheet_obj.cell(row = startRow, column = startCol)
                    semesterTitle = cell_obj.value
                    print(startRow)
                    if semesterTitle == None:
                        checkRow = startRow
                        while(checkRow < 358):
                            checkRow += 1
                            cell_obj = sheet_obj.cell(row = checkRow, column = startCol)
                            semesterTitle = cell_obj.value
                            if semesterTitle != None and "YEAR" in semesterTitle:
                                startRow = checkRow
                                break
                    print("blue")
                    print(startRow)
                    print(semesterTitle)
                    """
            flash('Semester Grade added!', category='success')
        ###
    print(pars)
    theUser = User.query.filter_by(matriculationNumber=pars).first()
    check = False
    if current_user.matriculationNumber == "admin":
        check = True
    return render_template("home.html", user=theUser, isAdmin=check, name=pars)


@views.route('/delete-note', methods=['POST'])
def delete_note():
    note = json.loads(request.data)
    noteId = note['noteId']
    note = Note.query.get(noteId)
    if note:
        db.session.delete(note)
        db.session.commit()

    return jsonify({})

@views.route('/delete-sg', methods=['POST'])
def delete_sg():
    sg = json.loads(request.data)
    sgId = sg['sgId']
    sg = SemesterGrade.query.get(sgId)
    if sg:
        db.session.delete(sg)
        db.session.commit()

    return jsonify({})